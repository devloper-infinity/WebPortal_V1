from pathlib import Path
import openpyxl


SOURCE = Path(r"C:\Users\ngk\Downloads\RCA Drop Down as per Taxonomy - Credit and Compliance V3.xlsx")
OUTPUT = Path(__file__).resolve().parent
def sql_text(value):
    return "N'" + str(value).replace("'", "''") + "'"


def unique_in_order(values):
    seen = set()
    result = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def emit_values(table, columns, rows):
    lines = [f"INSERT dbo.{table} ({', '.join(columns)}) VALUES"]
    formatted = []
    for row in rows:
        formatted.append("    (" + ", ".join(sql_text(v) if isinstance(v, str) else str(v) for v in row) + ")")
    lines.append(",\n".join(formatted) + ";")
    return "\n".join(lines)


workbook = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
TAXONOMY_SHEETS = [
    str(row[0]).strip()
    for row in workbook["Index"].iter_rows(values_only=True)
    if row[0] is not None and str(row[0]).strip().endswith("Taxonomy")
]
if len(TAXONOMY_SHEETS) != 9 or any(name not in workbook.sheetnames for name in TAXONOMY_SHEETS):
    raise ValueError("Index must contain the nine matching taxonomy sheet names.")

error_rows = [
    tuple("" if value is None else str(value).strip() for value in row[:9])
    for row in workbook["Error 1 to 4"].iter_rows(min_row=2, values_only=True)
]

et1_names = unique_in_order(row[0] for row in error_rows if row[0])
et1_id = {name: index + 1 for index, name in enumerate(et1_names)}
et2_pairs = unique_in_order((row[0], row[1]) for row in error_rows if row[0] and row[1])
et2_id = {pair: index + 1 for index, pair in enumerate(et2_pairs)}
et3_paths = unique_in_order((row[0], row[1], row[2]) for row in error_rows if row[0] and row[1] and row[2])
et4_names = unique_in_order(row[3] for row in error_rows if row[3])
et8_names = unique_in_order(row[7] for row in error_rows if row[7])
et9_names = unique_in_order(row[8] for row in error_rows if row[8])

master_sections = [
    "/* Generated from workbook sheet: Error 1 to 4. Do not hand-edit source values. */",
    "SET NOCOUNT ON;\nSET XACT_ABORT ON;\nBEGIN TRANSACTION;",
    emit_values("ErrorType1Master", ["ID", "Name", "IsActive", "DisplayOrder", "AddedBy", "AddedDate"],
                [(index + 1, name, 1, index + 1, 0, "GETDATE()") for index, name in enumerate(et1_names)]).replace("N'GETDATE()'", "GETDATE()"),
    emit_values("ErrorType2Master", ["ID", "ErrorType1ID", "Name", "IsActive", "DisplayOrder", "AddedBy", "AddedDate"],
                [(index + 1, et1_id[parent], name, 1, index + 1, 0, "GETDATE()") for index, (parent, name) in enumerate(et2_pairs)]).replace("N'GETDATE()'", "GETDATE()"),
    emit_values("ErrorType3Master", ["ID", "ErrorType2ID", "Name", "IsActive", "DisplayOrder", "AddedBy", "AddedDate"],
                [(index + 1, et2_id[(et1, et2)], name, 1, index + 1, 0, "GETDATE()") for index, (et1, et2, name) in enumerate(et3_paths)]).replace("N'GETDATE()'", "GETDATE()"),
]
for table, names in (("ErrorType4Master", et4_names), ("ErrorType8Master", et8_names), ("ErrorType9Master", et9_names)):
    master_sections.append(emit_values(table, ["ID", "Name", "IsActive", "DisplayOrder", "AddedBy", "AddedDate"],
                                      [(index + 1, name, 1, index + 1, 0, "GETDATE()") for index, name in enumerate(names)]).replace("N'GETDATE()'", "GETDATE()"))
master_sections.extend(["COMMIT TRANSACTION;", "GO", ""])
(OUTPUT / "003_Excel_Master_Data.sql").write_text("\n\n".join(master_sections), encoding="utf-8")

taxonomy_rows = []
et5_rows = []
et6_rows = []
et7_rows = []
et5_id = {}
et6_id = {}
next_et5 = next_et6 = next_et7 = 1

for taxonomy_index, sheet_name in enumerate(TAXONOMY_SHEETS, 1):
    taxonomy_rows.append((taxonomy_index, sheet_name, 1, taxonomy_index, 0, "GETDATE()"))
    rows = [
        tuple(str(value).strip() for value in row[:3])
        for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True)
        if all(value is not None and str(value).strip() for value in row[:3])
    ]
    for name5, name6, name7 in rows:
        key5 = (taxonomy_index, name5)
        if key5 not in et5_id:
            et5_id[key5] = next_et5
            et5_rows.append((next_et5, taxonomy_index, name5, 1, next_et5, 0, "GETDATE()"))
            next_et5 += 1
        key6 = (et5_id[key5], name6)
        if key6 not in et6_id:
            et6_id[key6] = next_et6
            et6_rows.append((next_et6, et5_id[key5], name6, 1, next_et6, 0, "GETDATE()"))
            next_et6 += 1
        et7_rows.append((next_et7, et6_id[key6], name7, 1, next_et7, 0, "GETDATE()"))
        next_et7 += 1

mapping_sections = [
    "/* Generated from Index and all nine taxonomy sheets. Internal feedbacks and Classification Selector are intentionally excluded. */",
    "SET NOCOUNT ON;\nSET XACT_ABORT ON;\nBEGIN TRANSACTION;",
]
for table, columns, rows in (
    ("TaxonomyMaster", ["ID", "Name", "IsActive", "DisplayOrder", "AddedBy", "AddedDate"], taxonomy_rows),
    ("ErrorType5Master", ["ID", "TaxonomyID", "Name", "IsActive", "DisplayOrder", "AddedBy", "AddedDate"], et5_rows),
    ("ErrorType6Master", ["ID", "ErrorType5ID", "Name", "IsActive", "DisplayOrder", "AddedBy", "AddedDate"], et6_rows),
    ("ErrorType7Master", ["ID", "ErrorType6ID", "Name", "IsActive", "DisplayOrder", "AddedBy", "AddedDate"], et7_rows),
):
    # Keep statements comfortably below SQL Server's 1,000-row VALUES limit.
    for start in range(0, len(rows), 1000):
        statement = emit_values(table, columns, rows[start:start + 1000]).replace("N'GETDATE()'", "GETDATE()")
        mapping_sections.append(statement)

mapping_sections.extend([
    "IF (SELECT COUNT(*) FROM dbo.TaxonomyMaster) <> 9 THROW 51000, 'Taxonomy count mismatch.', 1;",
    "IF (SELECT COUNT(*) FROM dbo.ErrorType5Master) <> 9 THROW 51000, 'ET5 count mismatch.', 1;",
    "IF (SELECT COUNT(*) FROM dbo.ErrorType6Master) <> 203 THROW 51000, 'ET6 count mismatch.', 1;",
    "IF (SELECT COUNT(*) FROM dbo.ErrorType7Master) <> 3789 THROW 51000, 'ET7 count mismatch.', 1;",
    "COMMIT TRANSACTION;",
    "GO",
    "",
])
(OUTPUT / "004_Taxonomy_Mappings.sql").write_text("\n\n".join(mapping_sections), encoding="utf-8")

print(f"ET1={len(et1_names)}, ET2={len(et2_pairs)}, ET3 relationships={len(et3_paths)}, ET4={len(et4_names)}, ET8={len(et8_names)}, ET9={len(et9_names)}")
print(f"Taxonomy={len(taxonomy_rows)}, ET5={len(et5_rows)}, ET6={len(et6_rows)}, ET7={len(et7_rows)}")
